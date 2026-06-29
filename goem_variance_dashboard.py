"""
GOEM 13-Week Rolling Supply Plan — Variance Dashboard
Cummins India
"""

import streamlit as st
import pandas as pd
import numpy as np
import datetime
import io
import re
import plotly.graph_objects as go
from openpyxl import load_workbook

st.set_page_config(page_title="GOEM Variance Dashboard", page_icon="⚙️", layout="wide")

CUMMINS_RED  = "#C00000"
DARK         = "#1A1A1A"
CHART_BLUE   = "#1565C0"
CHART_GREEN  = "#2E7D32"
CHART_ORANGE = "#E65100"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
html, body, [class*="css"] {{ font-family: 'Inter', sans-serif; }}
.stApp {{ background-color: #fff; }}
.block-container {{ padding-top: 0.6rem; padding-bottom: 2rem; max-width: 100%; }}
div[data-testid="stSidebar"] {{ background: #F4F4F4; border-right: 1px solid #E0E0E0; }}

.sidebar-brand {{
    background: {CUMMINS_RED}; color: white; padding: 12px 16px;
    border-radius: 6px; margin-bottom: 16px;
    font-weight: 700; font-size: 0.85rem;
}}
.sidebar-brand span {{ font-weight: 400; opacity: 0.8; font-size: 0.72rem; display: block; margin-top: 2px; }}
.sb-section {{
    font-size: 0.65rem; font-weight: 700; text-transform: uppercase;
    letter-spacing: 0.6px; color: #888; margin: 16px 0 6px;
}}

.page-header {{
    background: linear-gradient(135deg, {DARK} 0%, #3a3a3a 100%);
    color: white; padding: 16px 22px; border-radius: 8px; margin-bottom: 16px;
}}
.page-header h1 {{ font-size: 1.05rem; font-weight: 700; margin: 0; color: white; }}
.page-header p  {{ font-size: 0.72rem; margin: 4px 0 0; opacity: 0.7; }}

.mode-banner {{
    background: #E3F2FD; border-left: 4px solid #1565C0;
    padding: 7px 14px; border-radius: 0 4px 4px 0;
    font-size: 0.78rem; color: #0D47A1; margin-bottom: 14px;
}}

/* Comparison selector */
.cmp-selector {{
    background: #FAFAFA; border: 1px solid #E8E8E8; border-radius: 8px;
    padding: 14px 16px; margin-bottom: 16px;
}}
.cmp-label {{
    font-size: 0.65rem; font-weight: 700; text-transform: uppercase;
    letter-spacing: 0.5px; color: #888; margin-bottom: 6px;
}}
.cmp-arrow {{
    font-size: 1.4rem; color: {CUMMINS_RED}; font-weight: 700;
    text-align: center; line-height: 1; padding-top: 28px;
}}

/* View tabs */
.view-tabs {{
    display: flex; gap: 4px; margin-bottom: 16px;
    border-bottom: 2px solid #E0E0E0; padding-bottom: 0;
}}
.view-tab {{
    padding: 8px 16px; font-size: 0.78rem; font-weight: 600;
    border-radius: 6px 6px 0 0; cursor: pointer; color: #666;
    border: 1px solid transparent; border-bottom: none;
    transition: all 0.15s;
}}
.view-tab.active {{
    background: white; color: {CUMMINS_RED};
    border-color: #E0E0E0; border-bottom-color: white;
    margin-bottom: -2px;
}}

table.var-table {{
    width: 100%; border-collapse: collapse;
    font-size: 0.76rem; font-family: 'Inter', sans-serif;
}}
table.var-table thead tr:first-child th {{
    background: {DARK}; color: #fff; padding: 7px 10px;
    text-align: center; font-size: 0.69rem; font-weight: 600;
    border-right: 1px solid #333; white-space: nowrap;
}}
table.var-table thead tr:first-child th.left {{ text-align: left; }}
table.var-table thead tr:not(:first-child) th {{
    background: #2c2c2c; color: #ccc; padding: 4px 10px;
    font-size: 0.66rem; font-weight: 500;
    border-right: 1px solid #3a3a3a; white-space: nowrap; text-align: center;
}}
table.var-table tbody td {{
    padding: 5px 10px; border-bottom: 1px solid #F0F0F0;
    text-align: center; font-size: 0.74rem; color: {DARK};
    white-space: nowrap; border-right: 1px solid #F5F5F5;
}}
table.var-table tbody td.left {{ text-align: left; font-weight: 500; }}
table.var-table tbody tr:hover td {{ background: #FAFAFA !important; }}
table.var-table tbody tr.total-row td {{
    background: #F5F5F5 !important; font-weight: 700;
    border-top: 2px solid {CUMMINS_RED};
}}
.var-pos  {{ background-color: #E8F5E9 !important; color: #1B5E20; font-weight: 600; }}
.var-neg  {{ background-color: #FFEBEE !important; color: #B71C1C; font-weight: 600; }}
.var-zero {{ color: #BDBDBD; }}
.var-pos-pct {{ color: #2E7D32; font-size: 0.64rem; }}
.var-neg-pct {{ color: #C62828; font-size: 0.64rem; }}

.sec-hdr {{
    display: flex; align-items: center; gap: 10px; margin: 20px 0 10px;
}}
.sec-hdr-line {{ flex: 1; height: 1px; background: #E0E0E0; }}
.sec-hdr-text {{
    font-size: 0.7rem; font-weight: 700; text-transform: uppercase;
    letter-spacing: 0.8px; color: {CUMMINS_RED}; white-space: nowrap;
}}
</style>
""", unsafe_allow_html=True)

# ── helpers ───────────────────────────────────────────────────────────────────
def safe(v):
    return float(v) if v is not None and isinstance(v, (int, float)) else 0.0

def fmt(v):
    return "–" if v == 0 else str(int(round(v)))

def parse_date_cell(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, str):
        m = re.match(r'(\d{1,2})-(\d{2})-(\d{4})', v.strip())
        if m:
            day, mon, yr = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if mon == 0: mon = 6
            try: return datetime.date(yr, mon, day)
            except: pass
    return None

def var_td(a, b, show_pct=True):
    diff = a - b
    if diff == 0: return "var-zero", "–"
    sign = "+" if diff > 0 else ""
    cls  = "var-pos" if diff > 0 else "var-neg"
    pct_str = ""
    if show_pct and b != 0:
        pct = diff / b * 100
        pc  = "var-pos-pct" if pct > 0 else "var-neg-pct"
        pct_str = f'<br><span class="{pc}">{sign}{pct:.1f}%</span>'
    return cls, f'{sign}{int(round(diff))}{pct_str}'

def to_excel_bytes(dfs):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='xlsxwriter') as w:
        for name, df in dfs.items():
            if df is not None and not df.empty:
                df.to_excel(w, sheet_name=name[:31], index=False)
    return buf.getvalue()

def bucket_label(wk):
    if wk <= 6: return "0–6 Wks"
    elif wk <= 10: return "7–10 Wks"
    elif wk <= 13: return "11–13 Wks"
    else: return "14+ Wks"

BUCKET_ORDER = ["0–6 Wks", "7–10 Wks", "11–13 Wks", "14+ Wks"]

def section_header(text):
    st.markdown(f'''<div class="sec-hdr">
      <div class="sec-hdr-text">{text}</div>
      <div class="sec-hdr-line"></div>
    </div>''', unsafe_allow_html=True)

# ── parsing ───────────────────────────────────────────────────────────────────
def week_start(d):
    """Snap any date to the Monday of its week — canonical key for cross-sheet alignment."""
    return d - datetime.timedelta(days=d.weekday())

def get_true_week_cols(header, s, e):
    dated = [(i, parse_date_cell(v)) for i, v in enumerate(header) if i >= s and i <= e]
    dated = [(i, d) for i, d in dated if d is not None]
    # Snap to week_start, keep first occurrence per canonical week
    seen, weeks, prev_ws = {}, [], None
    for i, d in dated:
        ws = week_start(d)
        if ws not in seen:
            seen[ws] = i
            if prev_ws is None or (ws - prev_ws).days >= 6:
                weeks.append((i, ws))
                prev_ws = ws
    return weeks

def parse_customer_sheet(ws_rows, sheet_label):
    if len(ws_rows) < 3: return [], []
    header = ws_rows[2]
    dem_weeks = get_true_week_cols(header, 9, 31)
    sup_weeks = get_true_week_cols(header, 32, 54)
    dem_by_date = {d: i for i, d in dem_weeks}
    sup_by_date = {d: i for i, d in sup_weeks}
    all_dates = sorted(set(dem_by_date) | set(sup_by_date))
    records = []
    for row in ws_rows[3:]:
        if not row or (not row[0] and not row[3]): continue
        plant = str(row[0]).strip() if row[0] else ''
        kva   = safe(row[3]) if safe(row[3]) > 0 else safe(row[1])
        model = str(row[6]).strip() if row[6] else ''
        item  = str(row[7]).strip() if row[7] else ''
        scope = str(row[8]).strip() if row[8] else ''
        if not item or item in ('Shop Order', 'Coolpac Product No.', ''): continue
        if (sum(safe(row[i]) for i, d in dem_weeks if i < len(row)) == 0 and
            sum(safe(row[i]) for i, d in sup_weeks if i < len(row)) == 0): continue
        rec = {'plant': plant, 'kva': kva, 'model': model, 'item': item, 'scope': scope}
        for d in all_dates:
            dc = dem_by_date.get(d); sc = sup_by_date.get(d)
            rec[f'dem_{d}'] = safe(row[dc]) if dc and dc < len(row) else 0.0
            rec[f'sup_{d}'] = safe(row[sc]) if sc and sc < len(row) else 0.0
        records.append(rec)
    return records, all_dates

def parse_mds_sheet(ws_rows):
    if len(ws_rows) < 3: return [], []
    header = ws_rows[2]
    cutoff = datetime.date(2026, 6, 1)
    # Snap all dates to week_start, pick first col per canonical week, require >= 6 day gap
    seen = {}
    for i, v in enumerate(header):
        if not isinstance(v, datetime.datetime): continue
        d = v.date()
        if d < cutoff: continue
        ws = week_start(d)
        if ws not in seen:
            seen[ws] = i
    true_weeks, prev = [], None
    for ws_d, i in sorted(seen.items()):
        if prev is None or (ws_d - prev).days >= 6:
            true_weeks.append((i, ws_d))
        prev = ws_d
    date_col = {d: i for i, d in true_weeks}
    all_dates = sorted(date_col.keys())
    records = []
    for row in ws_rows[3:]:
        if not row or not row[1]: continue
        item  = str(row[1]).strip()
        model = str(row[5]).strip() if row[5] else ''
        kva   = safe(row[6]) if safe(row[6]) > 0 else safe(row[7])
        scope = str(row[8]).strip() if row[8] else ''
        if not item or item == 'Item No': continue
        rec = {'plant': str(row[0]).strip() if row[0] else '', 'kva': kva,
               'model': model, 'item': item, 'scope': scope}
        has = False
        for d in all_dates:
            v = safe(row[date_col[d]]) if date_col[d] < len(row) else 0.0
            rec[f'dem_{d}'] = v; rec[f'sup_{d}'] = 0.0
            if v: has = True
        if has: records.append(rec)
    return records, all_dates

def parse_plant_sp_sheet(ws_rows):
    if len(ws_rows) < 3: return [], []
    header = ws_rows[2]
    cutoff = datetime.date(2026, 6, 1)
    seen = {}
    for i, v in enumerate(header):
        if not isinstance(v, datetime.datetime): continue
        d = v.date()
        if d < cutoff: continue
        ws = week_start(d)
        if ws not in seen:
            seen[ws] = i
    true_weeks, prev = [], None
    for ws_d, i in sorted(seen.items()):
        if prev is None or (ws_d - prev).days >= 6:
            true_weeks.append((i, ws_d))
        prev = ws_d
    date_col = {d: i for i, d in true_weeks}
    all_dates = sorted(date_col.keys())
    records = []
    for row in ws_rows[3:]:
        if not row or not row[1]: continue
        item  = str(row[1]).strip()
        model = str(row[5]).strip() if row[5] else ''
        kva   = safe(row[6]) if safe(row[6]) > 0 else safe(row[7])
        scope = str(row[8]).strip() if row[8] else ''
        if not item or item == 'Item No': continue
        rec = {'plant': str(row[0]).strip() if row[0] else '', 'kva': kva,
               'model': model, 'item': item, 'scope': scope}
        has = False
        for d in all_dates:
            ci = date_col.get(d)
            v = safe(row[ci]) if ci and ci < len(row) else 0.0
            rec[f'sup_{d}'] = v; rec[f'dem_{d}'] = 0.0
            if v: has = True
        if has: records.append(rec)
    return records, all_dates

@st.cache_data(show_spinner=False)
def load_workbook_data(file_bytes):
    import io as _io
    wb = load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
    return {sh: [[c for c in row] for row in wb[sh].iter_rows(values_only=True)]
            for sh in wb.sheetnames}, wb.sheetnames

# ── table renderer ────────────────────────────────────────────────────────────
def render_table(rows, id_cols, id_labels, active_buckets, bucket_dates,
                 week_nums, show_pct, src_metric, ref_metric,
                 src_label, ref_label, start_date, weekly=False):
    n_id = len(id_cols)
    all_active_dates = sum(bucket_dates.values(), [])
    dates_13wk = [d for d in all_active_dates if week_nums[d] <= 13]

    TOTAL_STYLE  = 'border-left:3px solid #C00000;background:#fff8f8'
    TOTAL_STYLE2 = 'border-left:3px solid #C00000;background:#fff0f0'

    if weekly:
        # Header: bucket groups → weeks → [src, ref, var] per week + [src, ref, var] 13-wk total
        h = '<div style="overflow-x:auto"><table class="var-table"><thead><tr>'
        h += ''.join(f'<th class="left" rowspan="3">{l}</th>' for l in id_labels)
        for b in active_buckets:
            first_b = active_buckets[0]
            h += f'<th colspan="{len(bucket_dates[b])*3}" style="border-left:2px solid #444">{b}</th>'
        h += f'<th colspan="3" style="border-left:3px solid #C00000;background:#2a0000" rowspan="1">13-Wk Total</th>'
        h += '</tr><tr>'
        for b in active_buckets:
            for idx, d in enumerate(bucket_dates[b]):
                wk = week_nums[d]
                # First week of first bucket: show start_date, not the snapped date
                d_display = start_date if (b == active_buckets[0] and idx == 0) else d
                h += f'<th colspan="3" style="border-left:1px solid #444">Wk {wk}<br>{d_display.strftime("%d %b")}</th>'
        h += f'<th style="{TOTAL_STYLE}">{src_label[:8]}</th><th style="{TOTAL_STYLE}">{ref_label[:8]}</th><th style="{TOTAL_STYLE}">Variance</th>'
        h += '</tr><tr>'
        for b in active_buckets:
            for d in bucket_dates[b]:
                h += f'<th>{src_label[:8]}</th><th>{ref_label[:8]}</th><th>Variance</th>'
        h += f'<th style="{TOTAL_STYLE}"></th><th style="{TOTAL_STYLE}"></th><th style="{TOTAL_STYLE}"></th>'
        h += '</tr></thead><tbody>'

        tots = {d: [0.0, 0.0] for d in all_active_dates}
        grand = [0.0, 0.0]
        body = ''
        for r in rows:
            cells = ''.join(f'<td class="left">{r.get(k,"")}</td>' for k in id_cols)
            for b in active_buckets:
                for d in bucket_dates[b]:
                    a  = r.get(f'src_{src_metric}_{d}', 0)
                    bv = r.get(f'ref_{ref_metric}_{d}', 0)
                    tots[d][0] += a; tots[d][1] += bv
                    vc, vt = var_td(a, bv, show_pct)
                    cells += f'<td>{fmt(a)}</td><td>{fmt(bv)}</td><td class="{vc}">{vt}</td>'
            ta13 = sum(r.get(f'src_{src_metric}_{d}', 0) for d in dates_13wk)
            tb13 = sum(r.get(f'ref_{ref_metric}_{d}', 0) for d in dates_13wk)
            grand[0] += ta13; grand[1] += tb13
            vc13, vt13 = var_td(ta13, tb13, show_pct)
            cells += (f'<td style="{TOTAL_STYLE}"><strong>{fmt(ta13)}</strong></td>'
                      f'<td style="{TOTAL_STYLE}"><strong>{fmt(tb13)}</strong></td>'
                      f'<td class="{vc13}" style="{TOTAL_STYLE}"><strong>{vt13}</strong></td>')
            body += f'<tr>{cells}</tr>'

        tcells = f'<td class="left" colspan="{n_id}"><strong>TOTAL</strong></td>'
        for b in active_buckets:
            for d in bucket_dates[b]:
                ta, tb = tots[d]
                vc, vt = var_td(ta, tb, show_pct)
                tcells += f'<td><strong>{fmt(ta)}</strong></td><td><strong>{fmt(tb)}</strong></td><td class="{vc}"><strong>{vt}</strong></td>'
        vc13, vt13 = var_td(grand[0], grand[1], show_pct)
        tcells += (f'<td style="{TOTAL_STYLE2}"><strong>{fmt(grand[0])}</strong></td>'
                   f'<td style="{TOTAL_STYLE2}"><strong>{fmt(grand[1])}</strong></td>'
                   f'<td class="{vc13}" style="{TOTAL_STYLE2}"><strong>{vt13}</strong></td>')
        body += f'<tr class="total-row">{tcells}</tr>'
        return h + body + '</tbody></table></div>'

    else:
        # Bucket table: one col group per bucket + 13-wk total col at end
        h = '<div style="overflow-x:auto"><table class="var-table"><thead><tr>'
        h += ''.join(f'<th class="left" rowspan="2">{l}</th>' for l in id_labels)
        for b in active_buckets:
            # Show start_date as the range start for the first bucket
            first_b = active_buckets[0]
            d0_display = start_date if b == first_b else bucket_dates[b][0]
            d0 = d0_display.strftime("%d %b")
            d1 = bucket_dates[b][-1].strftime("%d %b")
            h += f'<th colspan="3" style="border-left:2px solid #444">{b}<br><span style="font-size:0.62rem;font-weight:400;opacity:0.8">{d0}–{d1}</span></th>'
        if dates_13wk:
            d0 = start_date.strftime("%d %b")
            d1 = dates_13wk[-1].strftime("%d %b")
            h += f'<th colspan="3" style="border-left:3px solid #C00000;background:#2a0000">13-Wk Total<br><span style="font-size:0.62rem;font-weight:400;opacity:0.8">{d0}–{d1}</span></th>'
        h += '</tr><tr>'
        for b in active_buckets:
            h += f'<th style="border-left:1px solid #3a3a3a">{src_label[:10]}</th><th>{ref_label[:10]}</th><th>Variance</th>'
        if dates_13wk:
            h += f'<th style="{TOTAL_STYLE}">{src_label[:10]}</th><th style="{TOTAL_STYLE}">{ref_label[:10]}</th><th style="{TOTAL_STYLE}">Variance</th>'
        h += '</tr></thead><tbody>'

        btots = {b: [0.0, 0.0] for b in active_buckets}
        grand = [0.0, 0.0]
        body = ''
        for r in rows:
            cells = ''.join(f'<td class="left">{r.get(k,"")}</td>' for k in id_cols)
            for b in active_buckets:
                a  = sum(r.get(f'src_{src_metric}_{d}', 0) for d in bucket_dates[b])
                bv = sum(r.get(f'ref_{ref_metric}_{d}', 0) for d in bucket_dates[b])
                btots[b][0] += a; btots[b][1] += bv
                vc, vt = var_td(a, bv, show_pct)
                cells += f'<td style="border-left:1px solid #eee">{fmt(a)}</td><td>{fmt(bv)}</td><td class="{vc}">{vt}</td>'
            if dates_13wk:
                ta13 = sum(r.get(f'src_{src_metric}_{d}', 0) for d in dates_13wk)
                tb13 = sum(r.get(f'ref_{ref_metric}_{d}', 0) for d in dates_13wk)
                grand[0] += ta13; grand[1] += tb13
                vc13, vt13 = var_td(ta13, tb13, show_pct)
                cells += (f'<td style="{TOTAL_STYLE}"><strong>{fmt(ta13)}</strong></td>'
                          f'<td style="{TOTAL_STYLE}"><strong>{fmt(tb13)}</strong></td>'
                          f'<td class="{vc13}" style="{TOTAL_STYLE}"><strong>{vt13}</strong></td>')
            body += f'<tr>{cells}</tr>'

        tcells = f'<td class="left" colspan="{n_id}"><strong>TOTAL</strong></td>'
        for b in active_buckets:
            ta, tb = btots[b]
            vc, vt = var_td(ta, tb, show_pct)
            tcells += f'<td style="border-left:1px solid #eee"><strong>{fmt(ta)}</strong></td><td><strong>{fmt(tb)}</strong></td><td class="{vc}"><strong>{vt}</strong></td>'
        if dates_13wk:
            vc13, vt13 = var_td(grand[0], grand[1], show_pct)
            tcells += (f'<td style="{TOTAL_STYLE2}"><strong>{fmt(grand[0])}</strong></td>'
                       f'<td style="{TOTAL_STYLE2}"><strong>{fmt(grand[1])}</strong></td>'
                       f'<td class="{vc13}" style="{TOTAL_STYLE2}"><strong>{vt13}</strong></td>')
        body += f'<tr class="total-row">{tcells}</tr>'
        return h + body + '</tbody></table></div>'


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown('''<div class="sidebar-brand">GOEM Variance Dashboard
      <span>Cummins India — Supply Planning</span>
    </div>''', unsafe_allow_html=True)
    uploaded = st.file_uploader("Upload Rolling Supply Plan (.xlsx)", type=["xlsx"],
                                label_visibility="collapsed")

if not uploaded:
    st.markdown('''<div class="page-header">
      <h1>GOEM 13-Week Rolling Supply Plan — Variance Dashboard</h1>
      <p>Upload the supply plan Excel file in the sidebar to begin</p>
    </div>''', unsafe_allow_html=True)
    st.info("Use the sidebar to upload your rolling supply plan .xlsx file.")
    st.stop()

raw_data, sheet_names = load_workbook_data(uploaded.read())

avail_customer = [s for s in ['SPL', 'PL', 'JL'] if s in sheet_names]
compare_sheet_actual = {s.strip(): s for s in ['Total', 'MDS', 'Plant SP '] if s in sheet_names}

# Parse all sheets upfront
parsed = {}
all_dates_set = set()

for sh in avail_customer:
    recs, dates = parse_customer_sheet(raw_data[sh], sh)
    parsed[sh] = (recs, dates); all_dates_set.update(dates)

for clean, actual in compare_sheet_actual.items():
    if clean == 'Total':
        recs, dates = parse_customer_sheet(raw_data[actual], 'Total')
    elif clean == 'MDS':
        recs, dates = parse_mds_sheet(raw_data[actual])
    elif clean == 'Plant SP':
        recs, dates = parse_plant_sp_sheet(raw_data[actual])
    else: continue
    parsed[clean] = (recs, dates); all_dates_set.update(dates)

all_dates = sorted(all_dates_set)
auto_start = min(all_dates) if all_dates else datetime.date.today()

# ── sidebar controls ──────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="sb-section">Analysis Window</div>', unsafe_allow_html=True)
    use_auto = st.checkbox("Auto-detect start date", value=True)
    if use_auto:
        start_date = auto_start
        st.caption(f"Start: **{start_date.strftime('%d %b %Y')}**")
    else:
        start_date = st.date_input("Start Date", value=auto_start, label_visibility="collapsed")

    st.markdown('<div class="sb-section">Display Options</div>', unsafe_allow_html=True)
    show_pct = st.toggle("Show % variance", value=True)

active_dates = [d for d in all_dates if d >= start_date]
week_nums    = {d: ((d - start_date).days // 7) + 1 for d in active_dates}
bucket_dates = {b: [] for b in BUCKET_ORDER}
for d in active_dates:
    bucket_dates[bucket_label(week_nums[d])].append(d)
active_buckets = [b for b in BUCKET_ORDER if bucket_dates[b]]

# ── SOURCE and REFERENCE selector (dynamic) ───────────────────────────────────
# Each side picks: SHEET + METRIC (Demand or Supply)
# Available sources: individual customer sheets + Total + MDS + Plant SP
CUST_DISPLAY = {'SPL': 'Sudhir Power', 'PL': 'Powerica', 'JL': 'Jakson'}
SHEET_OPTIONS = {}
for sh in avail_customer:
    SHEET_OPTIONS[sh] = f"{CUST_DISPLAY[sh]} ({sh})"
SHEET_OPTIONS['Total']    = 'Total GOEM'
SHEET_OPTIONS['MDS']      = 'MDS'
SHEET_OPTIONS['Plant SP'] = 'Plant SP'

avail_sheet_keys = [k for k in SHEET_OPTIONS if k in parsed]

METRIC_OPTIONS = {'dem': 'Demand', 'sup': 'Supply'}

# ── page header ───────────────────────────────────────────────────────────────
st.markdown('''<div class="page-header">
  <h1>GOEM 13-Week Rolling Supply Plan — Variance Dashboard</h1>
  <p>Select source and reference below to compare any combination of demand or supply across sheets</p>
</div>''', unsafe_allow_html=True)

# ── Comparison Selector ───────────────────────────────────────────────────────
st.markdown('<div class="cmp-label" style="font-size:0.72rem;font-weight:700;text-transform:uppercase;letter-spacing:0.5px;color:#555;margin-bottom:8px">Comparison Configuration</div>', unsafe_allow_html=True)

sel_col1, arrow_col, sel_col2, opt_col = st.columns([3, 0.6, 3, 2])

with sel_col1:
    st.markdown('<div class="cmp-label">Source</div>', unsafe_allow_html=True)
    src_sheet_keys  = st.multiselect("Source sheet(s)", avail_sheet_keys,
                                     default=[k for k in ['SPL','PL','JL'] if k in avail_sheet_keys],
                                     format_func=lambda x: SHEET_OPTIONS[x],
                                     label_visibility="collapsed")
    src_metric = st.radio("Source metric", list(METRIC_OPTIONS.keys()),
                          format_func=lambda x: METRIC_OPTIONS[x],
                          horizontal=True, key="src_metric", label_visibility="collapsed")

with arrow_col:
    st.markdown('<div style="padding-top:28px;text-align:center;font-size:1.6rem;color:#C00000;font-weight:700">→</div>', unsafe_allow_html=True)

with sel_col2:
    st.markdown('<div class="cmp-label">Reference</div>', unsafe_allow_html=True)
    ref_sheet_key = st.selectbox("Reference sheet", avail_sheet_keys,
                                 index=avail_sheet_keys.index('Total') if 'Total' in avail_sheet_keys else 0,
                                 format_func=lambda x: SHEET_OPTIONS[x],
                                 label_visibility="collapsed")
    ref_metric = st.radio("Reference metric", list(METRIC_OPTIONS.keys()),
                          format_func=lambda x: METRIC_OPTIONS[x],
                          horizontal=True, key="ref_metric", label_visibility="collapsed")

with opt_col:
    st.markdown('<div class="cmp-label">Filters</div>', unsafe_allow_html=True)
    only_var   = st.toggle("Variance items only", value=False)
    item_search = st.text_input("Search item", placeholder="e.g. SO60852", label_visibility="collapsed")

st.markdown("---")

if not src_sheet_keys:
    st.warning("Select at least one source sheet above.")
    st.stop()

src_label = " + ".join(SHEET_OPTIONS[k].split(" (")[0] for k in src_sheet_keys)
ref_label = SHEET_OPTIONS[ref_sheet_key].split(" (")[0]
src_metric_label = f"{src_label} {METRIC_OPTIONS[src_metric]}"
ref_metric_label = f"{ref_label} {METRIC_OPTIONS[ref_metric]}"

# Short labels for table column headers (keep readable, not truncated mid-word)
SHORT_NAMES = {'SPL': 'SPL', 'PL': 'PL', 'JL': 'JL',
               'Total': 'Total', 'MDS': 'MDS', 'Plant SP': 'PlantSP'}
if len(src_sheet_keys) == 1:
    src_short = SHORT_NAMES.get(src_sheet_keys[0], src_sheet_keys[0])
elif set(src_sheet_keys) == {'SPL', 'PL', 'JL'}:
    src_short = "GOEM"
else:
    src_short = "+".join(SHORT_NAMES.get(k, k) for k in src_sheet_keys)
ref_short = SHORT_NAMES.get(ref_sheet_key, ref_sheet_key)

src_col_label = f"{src_short} {METRIC_OPTIONS[src_metric][:3]}"   # e.g. "GOEM Dem"
ref_col_label = f"{ref_short} {METRIC_OPTIONS[ref_metric][:3]}"   # e.g. "PlantSP Sup"

# ── aggregate source data ─────────────────────────────────────────────────────
src_combined = []
for sh in src_sheet_keys:
    recs, _ = parsed[sh]
    src_combined.extend(recs)

src_agg = {}
for r in src_combined:
    key = (r['item'], r['model'], r['kva'], r['scope'])
    if key not in src_agg:
        src_agg[key] = {'item': r['item'], 'model': r['model'],
                        'kva': r['kva'], 'scope': r['scope']}
        for d in active_dates:
            src_agg[key][f'src_dem_{d}'] = 0.0
            src_agg[key][f'src_sup_{d}'] = 0.0
    for d in active_dates:
        src_agg[key][f'src_dem_{d}'] += r.get(f'dem_{d}', 0)
        src_agg[key][f'src_sup_{d}'] += r.get(f'sup_{d}', 0)

# ── aggregate reference data ──────────────────────────────────────────────────
ref_recs, _ = parsed.get(ref_sheet_key, ([], []))
ref_agg = {}
for r in ref_recs:
    key = (r['item'], r['model'], r['kva'], r['scope'])
    if key not in ref_agg:
        ref_agg[key] = {}
        for d in active_dates:
            ref_agg[key][f'ref_dem_{d}'] = 0.0
            ref_agg[key][f'ref_sup_{d}'] = 0.0
    for d in active_dates:
        ref_agg[key][f'ref_dem_{d}'] += r.get(f'dem_{d}', 0)
        ref_agg[key][f'ref_sup_{d}'] += r.get(f'sup_{d}', 0)

# ── merge ─────────────────────────────────────────────────────────────────────
merged = []
for key in set(src_agg) | set(ref_agg):
    row = {**(src_agg.get(key, {'item': key[0], 'model': key[1], 'kva': key[2], 'scope': key[3]})),
           **(ref_agg.get(key, {}))}
    for d in active_dates:
        row.setdefault(f'src_dem_{d}', 0.0); row.setdefault(f'src_sup_{d}', 0.0)
        row.setdefault(f'ref_dem_{d}', 0.0); row.setdefault(f'ref_sup_{d}', 0.0)
    merged.append(row)

merged = [r for r in merged if any(
    r.get(f'src_{src_metric}_{d}',0) + r.get(f'ref_{ref_metric}_{d}',0) > 0
    for d in active_dates)]
merged.sort(key=lambda x: (str(x.get('model','')), str(x.get('item',''))))

# ── filters ───────────────────────────────────────────────────────────────────
fc1, fc2 = st.columns([2, 2])
with fc1:
    models = sorted(set(r.get('model','') for r in merged if r.get('model')))
    sel_models = st.multiselect("Engine Model", models, key="f_model", placeholder="All models")
with fc2:
    kvas = sorted(set(r.get('kva',0) for r in merged if r.get('kva')))
    sel_kvas = st.multiselect("kVA Rating", [str(k) for k in kvas], key="f_kva", placeholder="All kVA")

if sel_models: merged = [r for r in merged if r.get('model','') in sel_models]
if sel_kvas:   merged = [r for r in merged if str(r.get('kva','')) in sel_kvas]
if item_search.strip():
    merged = [r for r in merged if item_search.strip().lower() in str(r.get('item','')).lower()]
if only_var:
    merged = [r for r in merged if any(
        r.get(f'src_{src_metric}_{d}',0) != r.get(f'ref_{ref_metric}_{d}',0)
        for d in active_dates)]

# ── totals ────────────────────────────────────────────────────────────────────
total_src = sum(r.get(f'src_{src_metric}_{d}',0) for r in merged for d in active_dates)
total_ref = sum(r.get(f'ref_{ref_metric}_{d}',0) for r in merged for d in active_dates)
variance  = total_src - total_ref

st.caption(f"**{len(merged)} items** · {src_metric_label} vs {ref_metric_label}")

# ── metric cards ──────────────────────────────────────────────────────────────
c1, c2, c3 = st.columns(3)
c1.metric(src_metric_label,  f"{int(total_src):,}")
c2.metric(ref_metric_label,  f"{int(total_ref):,}")
c3.metric("Variance",        f"{int(variance):+,}", delta=f"{int(variance):+,} units")

st.markdown("<br>", unsafe_allow_html=True)

# ── kVA aggregate helper ──────────────────────────────────────────────────────
def build_kva_rows():
    kva_agg = {}
    for r in merged:
        kv = r.get('kva', 0)
        if kv not in kva_agg:
            kva_agg[kv] = {'kva': kv}
            for d in active_dates:
                kva_agg[kv][f'src_dem_{d}'] = 0.0; kva_agg[kv][f'src_sup_{d}'] = 0.0
                kva_agg[kv][f'ref_dem_{d}'] = 0.0; kva_agg[kv][f'ref_sup_{d}'] = 0.0
        for d in active_dates:
            for m in ['dem','sup']:
                kva_agg[kv][f'src_{m}_{d}'] += r.get(f'src_{m}_{d}', 0)
                kva_agg[kv][f'ref_{m}_{d}'] += r.get(f'ref_{m}_{d}', 0)
    return sorted(kva_agg.values(), key=lambda x: float(x.get('kva',0) or 0))

def export_rows(rows, id_cols):
    dates_13wk = [d for d in sum(bucket_dates.values(), []) if week_nums[d] <= 13]
    out = []
    for r in rows:
        row_out = {k: r.get(k,'') for k in id_cols}
        for b in active_buckets:
            a  = sum(r.get(f'src_{src_metric}_{d}',0) for d in bucket_dates[b])
            bv = sum(r.get(f'ref_{ref_metric}_{d}',0) for d in bucket_dates[b])
            row_out[f'{b} {src_col_label}'] = int(a)
            row_out[f'{b} {ref_col_label}'] = int(bv)
            row_out[f'{b} Variance']        = int(a - bv)
        ta13 = sum(r.get(f'src_{src_metric}_{d}',0) for d in dates_13wk)
        tb13 = sum(r.get(f'ref_{ref_metric}_{d}',0) for d in dates_13wk)
        row_out[f'13-Wk Total {src_col_label}'] = int(ta13)
        row_out[f'13-Wk Total {ref_col_label}'] = int(tb13)
        row_out['13-Wk Total Variance']          = int(ta13 - tb13)
        out.append(row_out)
    return pd.DataFrame(out)

# ══════════════════════════════════════════════════════════════════════════════
# VIEW TABS
# ══════════════════════════════════════════════════════════════════════════════
tab_charts, tab_kva, tab_bucket, tab_weekly = st.tabs([
    "Charts & Overview", "kVA Summary", "Bucket Summary", "Week-by-Week Detail"
])

# ── TAB 1: CHARTS ─────────────────────────────────────────────────────────────
with tab_charts:
    ch1, ch2 = st.columns(2)

    with ch1:
        # Volume by bucket
        bcd = []
        for b in active_buckets:
            a  = sum(r.get(f'src_{src_metric}_{d}',0) for r in merged for d in bucket_dates[b])
            bv = sum(r.get(f'ref_{ref_metric}_{d}',0) for r in merged for d in bucket_dates[b])
            bcd.append({'b': b, 'src': a, 'ref': bv})
        fig1 = go.Figure()
        fig1.add_trace(go.Bar(name=src_col_label, x=[r['b'] for r in bcd],
                              y=[r['src'] for r in bcd], marker_color=CHART_BLUE, opacity=0.9))
        fig1.add_trace(go.Bar(name=ref_col_label, x=[r['b'] for r in bcd],
                              y=[r['ref'] for r in bcd], marker_color=CUMMINS_RED, opacity=0.7))
        fig1.update_layout(title=dict(text="Volume by Bucket", font=dict(size=13, color=DARK)),
            barmode='group', plot_bgcolor='white', paper_bgcolor='white',
            legend=dict(orientation='h', yanchor='bottom', y=1.01, x=1, xanchor='right', font=dict(size=10)),
            margin=dict(l=10,r=10,t=50,b=10), height=300,
            xaxis=dict(showgrid=False), yaxis=dict(gridcolor='#F0F0F0'))
        st.plotly_chart(fig1, use_container_width=True)

    with ch2:
        # Variance by bucket
        fig2 = go.Figure()
        for r in bcd:
            v = r['src'] - r['ref']
            fig2.add_trace(go.Bar(x=[r['b']], y=[v], name=r['b'], showlegend=False,
                marker_color=CHART_GREEN if v >= 0 else CUMMINS_RED, opacity=0.85,
                text=[f'{v:+,.0f}'], textposition='outside', textfont=dict(size=11)))
        fig2.update_layout(title=dict(text="Variance by Bucket (Source – Reference)", font=dict(size=13, color=DARK)),
            plot_bgcolor='white', paper_bgcolor='white',
            margin=dict(l=10,r=10,t=50,b=10), height=300,
            xaxis=dict(showgrid=False), yaxis=dict(gridcolor='#F0F0F0', zeroline=True, zerolinecolor='#999'))
        st.plotly_chart(fig2, use_container_width=True)

    ch3, ch4 = st.columns(2)

    with ch3:
        # Weekly trend
        wk_labels = [f"Wk{week_nums[d]} {d.strftime('%d%b')}" for d in active_dates]
        src_vals = [sum(r.get(f'src_{src_metric}_{d}',0) for r in merged) for d in active_dates]
        ref_vals = [sum(r.get(f'ref_{ref_metric}_{d}',0) for r in merged) for d in active_dates]
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(x=wk_labels, y=ref_vals, name=ref_col_label,
            line=dict(color='#90CAF9', width=1.5, dash='dash'), mode='lines'))
        fig3.add_trace(go.Scatter(x=wk_labels, y=src_vals, name=src_col_label,
            line=dict(color=CHART_BLUE, width=2.5), mode='lines+markers',
            marker=dict(size=5), fill='tonexty', fillcolor='rgba(21,101,192,0.08)'))
        fig3.update_layout(title=dict(text="Weekly Trend — Source vs Reference", font=dict(size=13, color=DARK)),
            plot_bgcolor='white', paper_bgcolor='white',
            legend=dict(orientation='h', yanchor='bottom', y=1.01, x=1, xanchor='right', font=dict(size=10)),
            margin=dict(l=10,r=10,t=50,b=10), height=300,
            xaxis=dict(showgrid=False, tickfont=dict(size=9), tickangle=-45),
            yaxis=dict(gridcolor='#F0F0F0'))
        st.plotly_chart(fig3, use_container_width=True)

    with ch4:
        # Variance by kVA
        kva_rows = build_kva_rows()
        kva_labels = [str(r['kva']) for r in kva_rows]
        kva_vars   = [sum(r.get(f'src_{src_metric}_{d}',0) - r.get(f'ref_{ref_metric}_{d}',0)
                         for d in active_dates) for r in kva_rows]
        fig4 = go.Figure(go.Bar(x=kva_labels, y=kva_vars,
            marker_color=[CHART_GREEN if v >= 0 else CUMMINS_RED for v in kva_vars], opacity=0.85,
            text=[f'{v:+,.0f}' for v in kva_vars], textposition='outside', textfont=dict(size=10)))
        fig4.update_layout(title=dict(text="Variance by kVA Rating", font=dict(size=13, color=DARK)),
            plot_bgcolor='white', paper_bgcolor='white',
            margin=dict(l=10,r=10,t=50,b=10), height=300,
            xaxis=dict(showgrid=False, tickfont=dict(size=10), title='kVA'),
            yaxis=dict(gridcolor='#F0F0F0', zeroline=True, zerolinecolor='#999'))
        st.plotly_chart(fig4, use_container_width=True)

# ── TAB 2: kVA SUMMARY ────────────────────────────────────────────────────────
with tab_kva:
    kva_rows = build_kva_rows()
    html = render_table(kva_rows, ['kva'], ['kVA'], active_buckets, bucket_dates,
                        week_nums, show_pct, src_metric, ref_metric,
                        src_col_label, ref_col_label, start_date, weekly=False)
    st.markdown(html, unsafe_allow_html=True)
    df_kva = export_rows(kva_rows, ['kva'])
    st.download_button("Export kVA Summary", data=to_excel_bytes({"kVA Summary": df_kva}),
                       file_name="goem_kva_variance.xlsx")

# ── TAB 3: BUCKET SUMMARY ─────────────────────────────────────────────────────
with tab_bucket:
    html = render_table(merged, ['item','model','kva','scope'],
                        ['Item No','Model','kVA','Scope'], active_buckets, bucket_dates,
                        week_nums, show_pct, src_metric, ref_metric,
                        src_col_label, ref_col_label, start_date, weekly=False)
    st.markdown(html, unsafe_allow_html=True)
    df_bucket = export_rows(merged, ['item','model','kva','scope'])
    df_bucket.rename(columns={'item':'Item No','model':'Model','kva':'kVA','scope':'Scope'}, inplace=True)
    st.download_button("Export Bucket Summary", data=to_excel_bytes({"Bucket Summary": df_bucket}),
                       file_name="goem_bucket_variance.xlsx")

# ── TAB 4: WEEKLY DETAIL ──────────────────────────────────────────────────────
with tab_weekly:
    html = render_table(merged, ['item','model','kva','scope'],
                        ['Item No','Model','kVA','Scope'], active_buckets, bucket_dates,
                        week_nums, show_pct, src_metric, ref_metric,
                        src_col_label, ref_col_label, start_date, weekly=True)
    st.markdown(html, unsafe_allow_html=True)
    # Weekly export
    exp_w = []
    for r in merged:
        row_out = {'Item No': r.get('item',''), 'Model': r.get('model',''),
                   'kVA': r.get('kva',''), 'Scope': r.get('scope','')}
        for d in active_dates:
            lbl = f"Wk{week_nums[d]} {d.strftime('%d%b')}"
            a  = r.get(f'src_{src_metric}_{d}', 0)
            bv = r.get(f'ref_{ref_metric}_{d}', 0)
            row_out[f'{lbl} {src_col_label}'] = int(a)
            row_out[f'{lbl} {ref_col_label}'] = int(bv)
            row_out[f'{lbl} Var']             = int(a - bv)
        exp_w.append(row_out)
    df_w = pd.DataFrame(exp_w)
    st.download_button("Export Weekly Detail",
                       data=to_excel_bytes({"Weekly Detail": df_w, "Bucket Summary": df_bucket,
                                            "kVA Summary": df_kva}),
                       file_name="goem_full_variance.xlsx")